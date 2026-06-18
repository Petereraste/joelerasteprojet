import os
import re
import tempfile
from collections import defaultdict
from flask import Flask, render_template, request, jsonify, send_file
import pdfplumber
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO

app = Flask(__name__)
app.config['MAX_CONTENT_LENGTH'] = 200 * 1024 * 1024


def parse_num(s):
    if s is None:
        return None
    s = str(s).strip().replace('\xa0', '').replace(' ', '').replace('(cid:3031)', '')
    try:
        return int(s)
    except (ValueError, TypeError):
        return None


def words_to_rows(words, y_tolerance=3):
    rows = defaultdict(list)
    for w in words:
        y = round(w['top'] / y_tolerance) * y_tolerance
        rows[y].append(w)
    result = []
    for y in sorted(rows.keys()):
        row_words = sorted(rows[y], key=lambda w: w['x0'])
        result.append((y, row_words))
    return result


def words_in_band(row_words, x_min, x_max):
    tokens = [w['text'] for w in row_words if x_min <= w['x0'] < x_max]
    return ''.join(tokens)


def extract_date_from_text(text):
    """Extract first DD/MM/YYYY date found in text, return as YYYY-MM-DD."""
    m = re.search(r'(\d{2})/(\d{2})/(\d{4})', text)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return None


MONTH_FR = {
    'jan': 1, 'fév': 2, 'fev': 2, 'mar': 3, 'avr': 4, 'mai': 5,
    'jun': 6, 'juin': 6, 'jul': 7, 'juil': 7, 'aoû': 8, 'aou': 8,
    'sep': 9, 'oct': 10, 'nov': 11, 'déc': 12, 'dec': 12
}


def parse_date_fr(s):
    """Parse date like '03-juin.-2026' or '03-juin-2026' to YYYY-MM-DD."""
    if not s:
        return None
    m = re.match(r'(\d{2})-([a-zA-ZéèêàûôîäëïüùæœÉÈÊÀÛÔÎÄËÏÜÙÆŒ]+)\.?-(\d{4})', s.strip())
    if m:
        day = int(m.group(1))
        month_str = m.group(2).lower()[:4].rstrip('.')
        # Try various prefix lengths
        month_num = None
        for length in [4, 3, 2]:
            key = month_str[:length]
            if key in MONTH_FR:
                month_num = MONTH_FR[key]
                break
        if month_num:
            return f"{m.group(3)}-{month_num:02d}-{day:02d}"
    # fallback DD-MM-YYYY
    m2 = re.match(r'(\d{2})-(\d{2})-(\d{4})', s.strip())
    if m2:
        return f"{m2.group(3)}-{m2.group(2)}-{m2.group(1)}"
    return None


def detect_document_type(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = pdf.pages[0].extract_text() or ''
    text_upper = text.upper()
    if 'SOLDES AUX COMPTES' in text_upper:
        return 'soldes'
    elif 'VALORISATION PORTEFEUILLE' in text_upper:
        return 'valorisation'
    elif 'RAPPORT QUOTIDIEN DES AFFECTATIONS' in text_upper or ('AFFECTATIONS' in text_upper and 'PRELIMINAIRE' in text_upper):
        return 'affectations'
    elif 'SITUATION CLIENT' in text_upper and 'MATRICULE CLIENT' in text_upper:
        return 'situation_client'
    elif 'RELEVE TITRES' in text_upper or 'RELEVÉ TITRES' in text_upper or ('HONNEUR DE VOUS ADRESSER' in text_upper and 'TITRES' in text_upper):
        return 'releve_titres'
    return 'unknown'


# ── Column boundaries ──────────────────────────────────────────────────
SAC_TITRE_MAX = 200
SAC_ISIN_MIN = 200
SAC_ISIN_MAX = 330
SAC_DISPO_MIN = 330
SAC_DISPO_MAX = 410
SAC_VENTE_MIN = 410
SAC_VENTE_MAX = 450
SAC_GELE_MIN = 450
SAC_GELE_MAX = 530
SAC_SOLDE_IND_MIN = 530

VP_COMPTE_MAX = 210
VP_NCOMPTE_MIN = 210
VP_NCOMPTE_MAX = 260
VP_TITRE_MIN = 260
VP_TITRE_MAX = 365
VP_COURS_MIN = 365
VP_COURS_MAX = 440
VP_BALANCE_MIN = 440
VP_BALANCE_MAX = 490
VP_VALO_MIN = 490

ISIN_RE = re.compile(r'^[A-Z]{2}[0-9]{10}$')


def parse_soldes_aux_comptes(pdf_path, source_file):
    accounts = {}
    current_adherent = None
    doc_date = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''

            if not doc_date:
                doc_date = extract_date_from_text(text)

            for line in text.split('\n'):
                m = re.match(r'Adh[eé]rent\s*-\s*(\d+)\s+(.+?)\s+(\d{2}/\d{2}/\d{4})', line)
                if m:
                    adh_id = m.group(1).strip()
                    adh_name = m.group(2).strip()
                    line_date = f"{m.group(3)[6:]}-{m.group(3)[3:5]}-{m.group(3)[:2]}"
                    current_adherent = adh_id
                    if adh_id not in accounts:
                        accounts[adh_id] = {
                            'adherent': adh_id,
                            'name': adh_name,
                            'date': line_date or doc_date,
                            'source_file': source_file,
                            'lines': []
                        }

            if not current_adherent:
                continue

            words = page.extract_words()
            rows = words_to_rows(words, y_tolerance=4)

            for y, row_words in rows:
                isin_words = [w for w in row_words
                              if SAC_ISIN_MIN <= w['x0'] < SAC_ISIN_MAX
                              and ISIN_RE.match(w['text'])]
                if not isin_words:
                    continue

                isin = isin_words[0]['text']
                titre_words = [w['text'] for w in row_words if w['x0'] < SAC_TITRE_MAX]
                titre = ' '.join(titre_words).strip()

                dispo = parse_num(words_in_band(row_words, SAC_DISPO_MIN, SAC_DISPO_MAX))
                vente = parse_num(words_in_band(row_words, SAC_VENTE_MIN, SAC_VENTE_MAX))
                gele = parse_num(words_in_band(row_words, SAC_GELE_MIN, SAC_GELE_MAX))
                solde_ind = parse_num(words_in_band(row_words, SAC_SOLDE_IND_MIN, 9999))

                accounts[current_adherent]['lines'].append({
                    'titre': titre,
                    'isin': isin,
                    'solde_disponible': dispo or 0,
                    'vente_en_attente': vente or 0,
                    'gele': gele or 0,
                    'nanti': 0,
                    'solde_indicatif': solde_ind or dispo or 0,
                    'source_file': source_file,
                })

    return list(accounts.values())


def parse_valorisation_portefeuille(pdf_path, source_file):
    accounts = {}
    doc_date = None

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ''
            words = page.extract_words()
            rows = words_to_rows(words, y_tolerance=4)

            if not doc_date:
                doc_date = extract_date_from_text(text)

            m_total = re.search(r'Ligne de portefeuille\s*:\s*\d+\s+([\d\s]+)', text)
            page_total = None
            if m_total:
                page_total = parse_num(m_total.group(1).replace(' ', ''))

            last_account = None
            for y, row_words in rows:
                n_compte = None
                compte_name_words = []
                for w in row_words:
                    m_nc = re.search(r'(\d{6})', w['text'])
                    if m_nc and w['x0'] < VP_TITRE_MIN:
                        n_compte = m_nc.group(1)
                        for pw in row_words:
                            if pw is w:
                                break
                            if pw['x0'] < VP_COMPTE_MAX:
                                compte_name_words.append(pw['text'])
                        break

                if not n_compte:
                    continue

                compte_name = ' '.join(compte_name_words).strip()
                if not compte_name:
                    for w in row_words:
                        m_nc = re.search(r'(\d{6})', w['text'])
                        if m_nc:
                            prefix = w['text'][:m_nc.start()].strip()
                            if prefix:
                                compte_name = prefix
                            break

                titre_words = [w['text'] for w in row_words
                               if VP_TITRE_MIN <= w['x0'] < VP_TITRE_MAX]
                titre = ' '.join(titre_words).strip()

                if not titre:
                    continue

                cours = parse_num(words_in_band(row_words, VP_COURS_MIN, VP_COURS_MAX))
                balance = parse_num(words_in_band(row_words, VP_BALANCE_MIN, VP_BALANCE_MAX))
                valo = parse_num(words_in_band(row_words, VP_VALO_MIN, 9999))

                if n_compte not in accounts:
                    accounts[n_compte] = {
                        'adherent': n_compte,
                        'name': compte_name,
                        'date': doc_date,
                        'source_file': source_file,
                        'total_valorisation': 0,
                        'lines': []
                    }
                accounts[n_compte]['lines'].append({
                    'titre': titre,
                    'cours_reference': cours,
                    'balance': balance or 0,
                    'valorisation': valo or 0,
                    'source_file': source_file,
                })
                last_account = n_compte

            if page_total and last_account:
                accounts[last_account]['total_valorisation'] = page_total

    for key in accounts:
        if not accounts[key]['total_valorisation']:
            accounts[key]['total_valorisation'] = sum(
                l['valorisation'] for l in accounts[key]['lines']
            )

    return list(accounts.values())


# ── Affectations Préliminaires column bands ────────────────────────────
AFF_TITRE_MIN = 37
AFF_TITRE_MAX = 136
AFF_IDAFFECT_MIN = 136
AFF_IDAFFECT_MAX = 183
AFF_TRANS_MIN = 183
AFF_TRANS_MAX = 223
AFF_NEG_MIN = 223
AFF_NEG_MAX = 329
AFF_ORDRE_MIN = 329
AFF_ORDRE_MAX = 369
AFF_COMPTE_MIN = 358
AFF_COMPTE_MAX = 412
AFF_QTE_MIN = 412
AFF_QTE_MAX = 463
AFF_COURS_MIN = 463
AFF_COURS_MAX = 518
AFF_VALEUR_MIN = 518


def parse_affectations_preliminaires(pdf_path, source_file):
    results = []

    with pdfplumber.open(pdf_path) as pdf:
        adherent = None
        date_seance = None
        date_reglement = None
        lines = []

        for page in pdf.pages:
            text = page.extract_text() or ''

            # Extract header fields
            m = re.search(r'Adh[eé]rent\s*:\s*(.+)', text)
            if m:
                adherent = m.group(1).strip()

            m = re.search(r'S[eé]ance de Bourse\s*:\s*(\d{2}-\S+-\d{4})', text)
            if m:
                date_seance = parse_date_fr(m.group(1))

            m = re.search(r'Date R[eè]glement\s*:\s*(\d{2}-\S+-\d{4})', text)
            if m:
                date_reglement = parse_date_fr(m.group(1))

            words = page.extract_words()
            rows = words_to_rows(words, y_tolerance=3)

            for y, row_words in rows:
                # Data row: has a 6-digit number in the Compte band
                compte_text = words_in_band(row_words, AFF_COMPTE_MIN, AFF_COMPTE_MAX)
                if not re.match(r'^\d{6}$', compte_text.strip()):
                    continue

                titre_words = [w['text'] for w in row_words
                               if AFF_TITRE_MIN <= w['x0'] < AFF_TITRE_MAX]
                titre = ' '.join(titre_words).strip()

                neg_words = [w['text'] for w in row_words
                             if AFF_NEG_MIN <= w['x0'] < AFF_NEG_MAX]
                negociateur = ' '.join(neg_words).strip()

                qte_raw = words_in_band(row_words, AFF_QTE_MIN, AFF_QTE_MAX)
                cours_raw = words_in_band(row_words, AFF_COURS_MIN, AFF_COURS_MAX)
                valeur_raw = words_in_band(row_words, AFF_VALEUR_MIN, 9999)

                # Handle negative numbers: "-60" may span as separate tokens
                def parse_signed(s):
                    s = s.strip().replace('\xa0', '').replace(' ', '').replace('(cid:3031)', '')
                    try:
                        return int(s)
                    except (ValueError, TypeError):
                        return None

                lines.append({
                    'titre': titre,
                    'compte': compte_text.strip(),
                    'quantite': parse_signed(qte_raw),
                    'cours': parse_signed(cours_raw),
                    'valeur': parse_signed(valeur_raw),
                    'negociateur': negociateur,
                })

    if lines:
        results.append({
            'adherent': adherent or '',
            'date': date_seance or '',
            'date_reglement': date_reglement or '',
            'source_file': source_file,
            'lines': lines,
        })

    return results


# ── Situation Client column bands ──────────────────────────────────────
SC_DEVISE_MIN = 10
SC_DEVISE_MAX = 30
SC_QTE_MIN = 100
SC_QTE_MAX = 138
SC_TITRE_MIN = 138
SC_TITRE_MAX = 298
SC_ISIN_MIN = 138
SC_ISIN_MAX = 200
SC_COURS_MOY_MIN = 298
SC_COURS_MOY_MAX = 376
SC_COURS_ACT_MIN = 376
SC_COURS_ACT_MAX = 462
SC_PMV_MIN = 462
SC_PMV_MAX = 527
SC_PMV_LAT_MIN = 527
SC_PMV_LAT_MAX = 615
SC_VALO_MIN = 615
SC_VALO_MAX = 713
SC_CV_XOF_MIN = 713
SC_CV_XOF_MAX = 788


def parse_quantite_sc(s):
    """Parse quantity strings like '192', '4,648', '42,646.00' to int."""
    if not s:
        return None
    s = s.strip().replace(',', '').replace(' ', '')
    try:
        return int(float(s))
    except (ValueError, TypeError):
        return None


def parse_cours_sc(s):
    """Parse cours like '10,000.00' to float."""
    if not s:
        return None
    s = s.strip().replace(',', '').replace(' ', '')
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def parse_situation_client(pdf_path, source_file):
    results = []

    with pdfplumber.open(pdf_path) as pdf:
        client_name = None
        matricule = None
        doc_date = None
        lines = []

        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text() or ''

            # Extract header fields from early pages
            if page_num < 3:
                m = re.search(r'Nom client\s*:\s*(.+)', text)
                if m:
                    client_name = m.group(1).strip()

                m = re.search(r'Matricule Client\s*:\s*(\d+)', text)
                if m:
                    matricule = m.group(1).strip()

                m = re.search(r'Date\s*:\s*(\d{2}-\d{2}-\d{4})', text)
                if m:
                    d = m.group(1)
                    doc_date = f"{d[6:]}-{d[3:5]}-{d[:2]}"

            # Data rows from page 3+
            if page_num < 2:
                continue

            words = page.extract_words()
            rows = words_to_rows(words, y_tolerance=3)

            for idx, (y, row_words) in enumerate(rows):
                # Data row: has "XOF" in devise band AND quantity in qte band
                devise_words = [w['text'] for w in row_words
                                if SC_DEVISE_MIN <= w['x0'] < SC_DEVISE_MAX]
                if not any(d == 'XOF' for d in devise_words):
                    continue

                qte_text = words_in_band(row_words, SC_QTE_MIN, SC_QTE_MAX)
                if not qte_text.strip():
                    continue

                titre_words = [w['text'] for w in row_words
                               if SC_TITRE_MIN <= w['x0'] < SC_TITRE_MAX]
                titre = ' '.join(titre_words).strip()

                cours_act_text = words_in_band(row_words, SC_COURS_ACT_MIN, SC_COURS_ACT_MAX)
                valo_text = words_in_band(row_words, SC_VALO_MIN, SC_VALO_MAX)
                cv_xof_text = words_in_band(row_words, SC_CV_XOF_MIN, SC_CV_XOF_MAX)

                # Look at next row for ISIN
                isin = None
                if idx + 1 < len(rows):
                    _, next_row_words = rows[idx + 1]
                    for w in next_row_words:
                        if SC_ISIN_MIN <= w['x0'] < SC_ISIN_MAX:
                            candidate = w['text'].strip()
                            if ISIN_RE.match(candidate) or re.match(r'^[A-Z]{2,4}[0-9A-Z]*$', candidate):
                                isin = candidate
                                break

                lines.append({
                    'titre': titre,
                    'isin': isin,
                    'quantite': parse_quantite_sc(qte_text),
                    'cours_actuel': parse_cours_sc(cours_act_text),
                    'valorisation_xof': parse_num(cv_xof_text) if cv_xof_text.strip() else parse_num(valo_text),
                    'source_file': source_file,
                })

    if lines:
        results.append({
            'client_name': client_name or '',
            'matricule': matricule or '',
            'date': doc_date or '',
            'source_file': source_file,
            'lines': lines,
        })

    return results


# ── Relevé Titres FCP column bands ─────────────────────────────────────
RT_TITRE_MIN = 30
RT_TITRE_MAX = 240
RT_QTE_MIN = 248
RT_QTE_MAX = 295
RT_COUT_MIN = 295
RT_COUT_MAX = 345
RT_VALEUR_MIN = 475
RT_VALEUR_MAX = 545


def parse_releve_titres(pdf_path, source_file):
    results = []

    with pdfplumber.open(pdf_path) as pdf:
        client = None
        compte = None
        doc_date = None
        lines = []
        current_section = None

        full_text_page0 = pdf.pages[0].extract_text() or ''

        # Extract header from first page text
        m = re.search(r'Client\s+N[°o]\s*:\s*(\d+)', full_text_page0)
        if m:
            client = m.group(1).strip()

        m = re.search(r'Compte\s+N[°o][:\s]+(\S+)', full_text_page0)
        if m:
            compte = m.group(1).strip()

        m = re.search(r'[eé]valu[eé]\s+au\s+(\d{2}/\d{2}/\d{4})', full_text_page0, re.IGNORECASE)
        if m:
            d = m.group(1)
            doc_date = f"{d[6:]}-{d[3:5]}-{d[:2]}"

        for page in pdf.pages:
            text = page.extract_text() or ''

            # Also try to find header on any page
            if not client:
                m = re.search(r'Client\s+N[°o]\s*:\s*(\d+)', text)
                if m:
                    client = m.group(1).strip()
            if not compte:
                m = re.search(r'Compte\s+N[°o][:\s]+(\S+)', text)
                if m:
                    compte = m.group(1).strip()
            if not doc_date:
                m = re.search(r'[eé]valu[eé]\s+au\s+(\d{2}/\d{2}/\d{4})', text, re.IGNORECASE)
                if m:
                    d = m.group(1)
                    doc_date = f"{d[6:]}-{d[3:5]}-{d[:2]}"

            words = page.extract_words()
            rows = words_to_rows(words, y_tolerance=3)

            for y, row_words in rows:
                titre_text = ' '.join(w['text'] for w in row_words if RT_TITRE_MIN <= w['x0'] < RT_TITRE_MAX).strip()
                qte_text = words_in_band(row_words, RT_QTE_MIN, RT_QTE_MAX).strip()

                # Detect section headers
                if titre_text in ('ACTIONS', 'OBLIGATIONS') and not qte_text:
                    current_section = titre_text
                    continue

                # Skip total/summary rows
                if titre_text in ('TOTAL', '') :
                    continue
                if re.match(r'^\d[\d\s]+\d+,\d+%', titre_text):
                    continue

                # Data row: titre band has text AND qte band has a number
                if not titre_text or not qte_text:
                    continue

                # Skip rows where qte looks non-numeric
                qte_clean = qte_text.replace(',', '').replace(' ', '').replace('\xa0', '')
                try:
                    quantite = int(qte_clean)
                except (ValueError, TypeError):
                    continue

                cout_raw = words_in_band(row_words, RT_COUT_MIN, RT_COUT_MAX).strip()
                valeur_raw = words_in_band(row_words, RT_VALEUR_MIN, RT_VALEUR_MAX).strip()

                # Parse cout_moyen as float
                def parse_float_num(s):
                    if not s:
                        return None
                    s = s.replace('\xa0', '').replace(' ', '').replace(',', '').replace('(cid:3031)', '')
                    try:
                        return float(s)
                    except (ValueError, TypeError):
                        return None

                lines.append({
                    'titre': titre_text,
                    'section': current_section or '',
                    'quantite': quantite,
                    'cout_moyen': parse_float_num(cout_raw),
                    'valeur_fcfa': parse_num(valeur_raw),
                    'source_file': source_file,
                })

    if lines:
        results.append({
            'client': client or '',
            'compte': compte or '',
            'date': doc_date or '',
            'source_file': source_file,
            'lines': lines,
        })

    return results


@app.route('/')
def index():
    return render_template('index.html')


@app.route('/api/parse', methods=['POST'])
def parse_pdf():
    if 'files' not in request.files:
        return jsonify({'error': 'No files provided'}), 400

    files = request.files.getlist('files')
    results = {
        'soldes': [],
        'valorisation': [],
        'affectations': [],
        'situation_client': [],
        'releve_titres': [],
        'errors': []
    }

    for f in files:
        if not f.filename.lower().endswith('.pdf'):
            results['errors'].append(f'{f.filename}: not a PDF')
            continue

        with tempfile.NamedTemporaryFile(suffix='.pdf', delete=False) as tmp:
            f.save(tmp.name)
            tmp_path = tmp.name

        try:
            doc_type = detect_document_type(tmp_path)
            if doc_type == 'soldes':
                accounts = parse_soldes_aux_comptes(tmp_path, f.filename)
                results['soldes'].extend(accounts)
            elif doc_type == 'valorisation':
                accounts = parse_valorisation_portefeuille(tmp_path, f.filename)
                results['valorisation'].extend(accounts)
            elif doc_type == 'affectations':
                accounts = parse_affectations_preliminaires(tmp_path, f.filename)
                results['affectations'].extend(accounts)
            elif doc_type == 'situation_client':
                accounts = parse_situation_client(tmp_path, f.filename)
                results['situation_client'].extend(accounts)
            elif doc_type == 'releve_titres':
                accounts = parse_releve_titres(tmp_path, f.filename)
                results['releve_titres'].extend(accounts)
            else:
                results['errors'].append(f'{f.filename}: type de document non reconnu')
        except Exception as e:
            import traceback
            results['errors'].append(f'{f.filename}: {str(e)}\n{traceback.format_exc()}')
        finally:
            os.unlink(tmp_path)

    return jsonify(results)


@app.route('/api/export', methods=['POST'])
def export_excel():
    data = request.get_json()
    if not data:
        return jsonify({'error': 'No data'}), 400

    # Optional date filter
    date_filter = data.get('date_filter')

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def hfill(color):
        return PatternFill(start_color=color, end_color=color, fill_type='solid')

    header_fill = hfill('003366')
    alt_fill = hfill('E8F0FE')
    total_fill = hfill('FFF2CC')
    ok_fill = hfill('D4EDDA')
    err_fill = hfill('F8D7DA')

    thin = Side(style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    num_fmt = '#,##0'

    def set_header(cell):
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True, size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    def set_cell(cell, row_i, is_num=False):
        cell.fill = alt_fill if row_i % 2 == 0 else PatternFill()
        cell.border = border
        cell.alignment = Alignment(vertical='center',
                                   horizontal='right' if is_num else 'left')
        if is_num:
            cell.number_format = num_fmt

    def matches_date(acc):
        if not date_filter:
            return True
        return acc.get('date') == date_filter

    # ── Soldes aux Comptes ────────────────────────────────────────────────
    soldes = [a for a in data.get('soldes', []) if matches_date(a)]
    if soldes:
        ws = wb.create_sheet('Soldes aux Comptes')
        cols = ['Date', 'Adhérent', 'Nom du Compte', 'Titre', 'Code ISIN',
                'Solde Disponible', 'Vente en Attente', 'Gelé', 'Nanti',
                'Solde Indicatif', 'Fichier Source']
        widths = [12, 12, 36, 40, 16, 18, 18, 10, 10, 18, 50]
        for c, (h, w) in enumerate(zip(cols, widths), 1):
            set_header(ws.cell(row=1, column=c, value=h))
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'
        row = 2
        for acc in soldes:
            for i, l in enumerate(acc.get('lines', [])):
                vals = [acc.get('date', ''), acc['adherent'], acc['name'],
                        l['titre'], l['isin'],
                        l['solde_disponible'], l['vente_en_attente'],
                        l['gele'], l['nanti'], l['solde_indicatif'],
                        l.get('source_file', acc.get('source_file', ''))]
                for c, v in enumerate(vals, 1):
                    set_cell(ws.cell(row=row, column=c, value=v), i, 6 <= c <= 10)
                row += 1

    # ── Valorisation Portefeuille ─────────────────────────────────────────
    valorisations = [a for a in data.get('valorisation', []) if matches_date(a)]
    if valorisations:
        ws = wb.create_sheet('Valorisation Portefeuille')
        cols = ['Date', 'N° Compte', 'Nom du Compte', 'Titre',
                'Cours Référence', 'Balance', 'Valorisation', 'Fichier Source']
        widths = [12, 12, 36, 40, 16, 14, 22, 50]
        for c, (h, w) in enumerate(zip(cols, widths), 1):
            set_header(ws.cell(row=1, column=c, value=h))
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'
        row = 2
        for acc in valorisations:
            for i, l in enumerate(acc.get('lines', [])):
                vals = [acc.get('date', ''), acc['adherent'], acc['name'],
                        l['titre'], l['cours_reference'], l['balance'],
                        l['valorisation'],
                        l.get('source_file', acc.get('source_file', ''))]
                for c, v in enumerate(vals, 1):
                    set_cell(ws.cell(row=row, column=c, value=v), i, 5 <= c <= 7)
                row += 1
            for c in range(1, 9):
                cell = ws.cell(row=row, column=c)
                cell.fill = total_fill
                cell.border = border
                cell.font = Font(bold=True)
            ws.cell(row=row, column=5, value='TOTAL').alignment = Alignment(horizontal='right')
            ws.cell(row=row, column=6,
                    value=sum(l['balance'] for l in acc.get('lines', []))).number_format = num_fmt
            ws.cell(row=row, column=7,
                    value=acc['total_valorisation']).number_format = num_fmt
            row += 2

    # ── Réconciliation ────────────────────────────────────────────────────
    if soldes and valorisations:
        ws = wb.create_sheet('Réconciliation')
        cols = ['Date', 'Adhérent', 'Nom du Compte', 'Titre',
                'Balance DCBR', 'Balance BTCC', 'Écart', 'Statut',
                'Fichier DCBR', 'Fichier BTCC']
        widths = [12, 12, 36, 40, 16, 16, 14, 12, 50, 50]
        for c, (h, w) in enumerate(zip(cols, widths), 1):
            set_header(ws.cell(row=1, column=c, value=h))
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'

        soldes_idx = {}
        for acc in soldes:
            adh = acc['adherent']
            date = acc.get('date', '')
            key = (adh, date)
            soldes_idx[key] = {}
            for l in acc.get('lines', []):
                k = l['titre'].lower().strip()[:12]
                soldes_idx[key][k] = {
                    'solde': l['solde_indicatif'],
                    'source': l.get('source_file', acc.get('source_file', ''))
                }

        row = 2
        for acc in valorisations:
            adh = acc['adherent']
            date = acc.get('date', '')
            key = (adh, date)
            for l in acc.get('lines', []):
                titre_key = l['titre'].lower().strip()[:12]
                balance_dcbr = None
                src_dcbr = ''
                if key in soldes_idx:
                    for k, v in soldes_idx[key].items():
                        if k[:8] == titre_key[:8]:
                            balance_dcbr = v['solde']
                            src_dcbr = v['source']
                            break

                balance_btcc = l['balance']
                if balance_dcbr is not None:
                    ecart = balance_btcc - balance_dcbr
                    statut = 'OK' if ecart == 0 else 'ÉCART'
                else:
                    ecart = None
                    statut = 'N/A'

                ok_f = hfill('D4EDDA')
                err_f = hfill('F8D7DA')
                fill = ok_f if statut == 'OK' else err_f if statut == 'ÉCART' else PatternFill()
                vals = [date, adh, acc['name'], l['titre'],
                        balance_dcbr, balance_btcc, ecart, statut,
                        src_dcbr, l.get('source_file', acc.get('source_file', ''))]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(row=row, column=c, value=v)
                    cell.fill = fill
                    cell.border = border
                    cell.alignment = Alignment(
                        vertical='center',
                        horizontal='right' if 5 <= c <= 7 else 'center' if c == 8 else 'left')
                    if c in (5, 6, 7):
                        cell.number_format = num_fmt
                    if c == 8:
                        cell.font = Font(bold=True,
                                         color=('276749' if statut == 'OK' else
                                                'C53030' if statut == 'ÉCART' else '4A5568'))
                row += 1

    # ── Affectations Préliminaires ────────────────────────────────────────
    affectations = [a for a in data.get('affectations', []) if matches_date(a)]
    if affectations:
        ws = wb.create_sheet('Affectations Préliminaires')
        cols = ['Date', 'Date Règlement', 'Adhérent', 'Titre', 'Compte',
                'Quantité', 'Cours', 'Valeur', 'Négociateur', 'Fichier Source']
        widths = [14, 14, 30, 40, 12, 12, 12, 18, 30, 50]
        for c, (h, w) in enumerate(zip(cols, widths), 1):
            set_header(ws.cell(row=1, column=c, value=h))
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'
        row = 2
        for acc in affectations:
            for i, l in enumerate(acc.get('lines', [])):
                vals = [
                    acc.get('date', ''),
                    acc.get('date_reglement', ''),
                    acc.get('adherent', ''),
                    l.get('titre', ''),
                    l.get('compte', ''),
                    l.get('quantite'),
                    l.get('cours'),
                    l.get('valeur'),
                    l.get('negociateur', ''),
                    l.get('source_file', acc.get('source_file', '')),
                ]
                for c, v in enumerate(vals, 1):
                    set_cell(ws.cell(row=row, column=c, value=v), i, 6 <= c <= 8)
                row += 1

    # ── Situation Client FCP ──────────────────────────────────────────────
    situation_clients = [a for a in data.get('situation_client', []) if matches_date(a)]
    if situation_clients:
        ws = wb.create_sheet('Situation Client FCP')
        cols = ['Date', 'Client', 'Matricule', 'Titre', 'ISIN',
                'Quantité', 'Cours Actuel', 'Valorisation XOF', 'Fichier Source']
        widths = [14, 36, 14, 40, 16, 12, 16, 20, 50]
        for c, (h, w) in enumerate(zip(cols, widths), 1):
            set_header(ws.cell(row=1, column=c, value=h))
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'
        row = 2
        for acc in situation_clients:
            for i, l in enumerate(acc.get('lines', [])):
                vals = [
                    acc.get('date', ''),
                    acc.get('client_name', ''),
                    acc.get('matricule', ''),
                    l.get('titre', ''),
                    l.get('isin', ''),
                    l.get('quantite'),
                    l.get('cours_actuel'),
                    l.get('valorisation_xof'),
                    l.get('source_file', acc.get('source_file', '')),
                ]
                for c, v in enumerate(vals, 1):
                    set_cell(ws.cell(row=row, column=c, value=v), i, c in (6, 7, 8))
                row += 1

    # ── Relevé Titres FCP ─────────────────────────────────────────────────
    releve_titres = [a for a in data.get('releve_titres', []) if matches_date(a)]
    if releve_titres:
        ws = wb.create_sheet('Relevé Titres FCP')
        cols = ['Date', 'Client', 'Compte', 'Section', 'Titre',
                'Quantité', 'Coût Moyen', 'Valeur FCFA', 'Fichier Source']
        widths = [14, 20, 16, 14, 40, 12, 18, 20, 50]
        for c, (h, w) in enumerate(zip(cols, widths), 1):
            set_header(ws.cell(row=1, column=c, value=h))
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.row_dimensions[1].height = 30
        ws.freeze_panes = 'A2'
        row = 2
        for acc in releve_titres:
            for i, l in enumerate(acc.get('lines', [])):
                vals = [
                    acc.get('date', ''),
                    acc.get('client', ''),
                    acc.get('compte', ''),
                    l.get('section', ''),
                    l.get('titre', ''),
                    l.get('quantite'),
                    l.get('cout_moyen'),
                    l.get('valeur_fcfa'),
                    l.get('source_file', acc.get('source_file', '')),
                ]
                for c, v in enumerate(vals, 1):
                    set_cell(ws.cell(row=row, column=c, value=v), i, c in (6, 7, 8))
                row += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Build filename with date if filtered
    fname = f'releves_{date_filter}.xlsx' if date_filter else 'releves_dcbr_btcc.xlsx'
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname
    )


if __name__ == '__main__':
    app.run(debug=True, port=5000)
